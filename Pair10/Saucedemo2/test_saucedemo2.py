# Generated by Selenium IDE
import pytest
import time
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import openpyxl
from pathlib import Path
from constants import *



class Test_Saucedemo2:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(BASE_DOMAIN_URL)
    def element_visible(self,id,value):
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((id, value)))
    def teardown_method(self):
        self.driver.quit()



# -standard_user girişi
    def test_Login(self):
        self.driver.get(BASE_DOMAIN_URL)
        self.driver.maximize_window()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, USERNAME)))
        self.driver.find_element(By.CSS_SELECTOR, USERNAME).send_keys(USERNAME_ID)
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, PASSWORD)))
        self.driver.find_element(By.CSS_SELECTOR, PASSWORD).send_keys(PASSWORD_ID)
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, LOGIN_BUTTON)))
        self.driver.find_element(By.CSS_SELECTOR, LOGIN_BUTTON).click()



# -excel
    def readUserDataFromExcel():
      excelFile = openpyxl.load_workbook("data/productNames.xlsx")
      selectedSheet = excelFile["Sheet1"]
      rows = selectedSheet.max_row
      data = []
      for i in range(2,rows+1):
        productname = selectedSheet.cell(i,1).value
        data.append(productname)
      return data



# -locked_out_user ile giriş yapıldığında verilen uyarı mesajının doğrulanması.
    def test_lockedoutuser(self):
        self.driver.get(BASE_DOMAIN_URL)
        self.driver.maximize_window()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, USERNAME)))
        self.driver.find_element(By.CSS_SELECTOR, USERNAME).send_keys(WRONG_USERNAME_ID)
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, PASSWORD)))
        self.driver.find_element(By.CSS_SELECTOR, PASSWORD).send_keys(PASSWORD_ID)
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, LOGIN_BUTTON)))
        self.driver.find_element(By.CSS_SELECTOR, LOGIN_BUTTON).click()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, ERROR_TEXT)))
        assert self.driver.find_element(By.CSS_SELECTOR, ERROR_TEXT).text == ERROR_TEXT_MESSAGE
  


# -normal giriş yapılıp sonrasında ürünlerin sayısının doğrulanması
    def test_productcount(self):
        self.test_Login()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CLASS_NAME, INVENTORY_ITEM)))
        products = self.driver.find_elements(By.CLASS_NAME, INVENTORY_ITEM)
        numberOfProducts = len(products)
        assert numberOfProducts == 6



# -ürünlerin isimlerinin excel dosyalarındaki isimlerle uyuşması
    def test_checkingNames(self):
        excelFile = openpyxl.load_workbook(PRODUCT_NAMES)
        selectedSheet = excelFile[FIRST_PAGE]
        rows = selectedSheet.max_row
        data = []
        for i in range(2,rows+1):
            productname = selectedSheet.cell(i,1).value
            data.append(productname)

        self.test_Login()
        products = self.driver.find_elements(By.CLASS_NAME, INVENTORY_ITEM_NAME)
        pnames = []
        
        for i in range(len(products)):
            product = products[i]
            productNamesText = product.text
            pnames.append(productNamesText)
        assert pnames[0] == data[0]



# -ürünlerin z'den a ya sıralanma fonksiyonun test edilmesi
    def test_productListZ2A(self):
        self.test_Login()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, PRODUCT_SORT)))
        self.driver.find_element(By.CSS_SELECTOR, PRODUCT_SORT).click()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.XPATH, PRODUCT_SORT_Z2A)))
        self.driver.find_element(By.XPATH, PRODUCT_SORT_Z2A).click()



# -ürünlerin düşük fiyattan yüksek fiyata sıralanma fonksiyonun test edilmesi
    def test_productListLow2High(self):
        self.test_Login()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, PRODUCT_SORT)))
        self.driver.find_element(By.CSS_SELECTOR, PRODUCT_SORT).click()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.XPATH, PRODUCT_SORT_L2H)))
        self.driver.find_element(By.XPATH, PRODUCT_SORT_L2H).click()
  


# -bir excel dosyasında ismi geçen ürünlerin sepete eklenmesi fonksiyonu testi
    def test_addBasketFromExcel(self):
        self.test_Login()
        excelFile = openpyxl.load_workbook(PRODUCTS_ON_BASKET)
        selectedSheet = excelFile[FIRST_PAGE]
        rows = selectedSheet.max_row
        data = []
        for i in range(2,rows+1):
            productName = selectedSheet.cell(i,1).value
            data.append(productName)

        if productName == FIRST_PRODUCT: 
            addBasket = self.driver.find_element(By.ID, FIRST_PRODUCT_ID)
            addBasket.click()
            basket = self.driver.find_elements(By.CLASS_NAME,BASKET_LINK)
            pOnBasket = len(basket)
            assert pOnBasket > 0

        elif productName == SECOND_PRODUCT: 
            addBasket = self.driver.find_element(By.ID, SECOND_PRODUCT_ID)
            addBasket.click()
            basket = self.driver.find_elements(By.CLASS_NAME,BASKET_LINK)
            pOnBasket = len(basket)
            assert pOnBasket > 0



# -sepete eklenen ürünlerin sepet sayfasında doğru bir şekilde görünmesi testi
    def test_productadd(self):
        self.test_Login()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CLASS_NAME, INVENTORY_ITEM_NAME)))
        self.driver.find_element(By.ID, FIRST_PRODUCT_ID).click()
        self.driver.find_element(By.CLASS_NAME, BASKET_LINK).click()
        addedBasket = self.driver.find_elements(By.CLASS_NAME,CHECK_BASKET_PRODUCTS)
        assert len(addedBasket) > 0



# -sepetten kaldırılan ürünün sepet ekranından kaldırılma testi
    def test_productremove(self):
        self.test_Login()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.CLASS_NAME, INVENTORY_ITEM_NAME)))
        self.driver.find_element(By.ID, FIRST_PRODUCT_ID).click()
        self.driver.find_element(By.CLASS_NAME, BASKET_LINK).click()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.XPATH, REMOVE_PRODUCT)))
        self.driver.find_element(By.XPATH, REMOVE_PRODUCT).click()
        removedBasket = self.driver.find_elements(By.CLASS_NAME,REMOVED_ITEM)
        assert len(removedBasket) > 0
